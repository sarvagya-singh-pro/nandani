import tensorflow as tf 
import numpy as np
import json

from flask import Flask, jsonify,request
from tensorflow import keras
import matplotlib.pyplot as plt
import math
from flask_cors import CORS, cross_origin
from sklearn.ensemble import RandomForestClassifier
from sklearn.metrics import accuracy_score
import pandas as pd



import pandas as pd
import openpyxl
# Import required libraries
import pandas as pd
import os
import numpy as np
import matplotlib.pyplot as plt


import sklearn

# Import necessary modules
from sklearn.model_selection import train_test_split
from sklearn.metrics import mean_squared_error
from math import sqrt

# Keras specific
import keras
from keras.models import Sequential
from keras.layers import Dense
from keras.utils import to_categorical


ws = openpyxl.load_workbook("Animal_Disease_Data.xlsx")
ws = ws.active
app = Flask(__name__)
symptoms = ["painless lumps",
"swelling in limb",
"sweats",
"loss of appetite"
,"swelling in extremities"
,"depression"
,"sores on hooves"
,"crackling sound"
,"swelling in muscle"
,"difficulty walking"
,"lameness"
,"fatigue"
,"blisters on hooves"
,"chills"
,"blisters on gums"
,"blisters on mouth"
,"swelling in abdomen"
,"chest discomfort"
,"swelling in neck"
,"shortness of breath"
,"sores on tongue"
,"blisters on tongue"
,"sores on gums"
,"sores on mouth"]

s_d = {}
count = 2
for symptom in symptoms:
  s_d[symptom] = count
  count+=1

print(s_d)

x_test = []
y_test_anthrax = []
y_test_blackleg = []
y_test_food_and_mouth_disease = []
y_test_lumpy_virus = []
y_test_pneumonia = []

for i in range(2,ws.max_row):
  temp = [0 for j in range(26)]
  temp[0] = (ws.cell(row = i, column = 1).value - 1)/14
  temp[1] = (ws.cell(row = i, column = 2).value - 100)/5
  temp[s_d[ws.cell(row = i, column = 3).value]] = 1
  temp[s_d[ws.cell(row = i, column = 4).value]] = 1
  temp[s_d[ws.cell(row = i, column = 5).value]] = 1
  x_test.append(temp)
  y_test_anthrax.append(ws.cell(row = i, column = 6).value == "anthrax")
  y_test_blackleg.append(ws.cell(row = i, column = 6).value == "blackleg")
  y_test_food_and_mouth_disease.append(ws.cell(row = i, column = 6).value == "foot and mouth")
  y_test_lumpy_virus.append(ws.cell(row = i, column = 6).value == "lumpy virus")
  y_test_pneumonia.append(ws.cell(row = i, column = 6).value == "pneumonia")

x_test = np.array(x_test)
y_test_anthrax = np.array(y_test_anthrax).reshape(len(y_test_anthrax), 1)
y_test_blackleg = np.array(y_test_blackleg).reshape(len(y_test_anthrax), 1)
y_test_food_and_mouth_disease =np.array(y_test_food_and_mouth_disease).reshape(len(y_test_anthrax), 1)
y_test_lumpy_virus =np.array(y_test_lumpy_virus).reshape(len(y_test_anthrax), 1)
y_test_pneumonia =  np.array(y_test_pneumonia).reshape(len(y_test_anthrax), 1)

y_test_anthrax_ = []
y_test_blackleg_ = []
y_test_food_and_mouth_disease_ = []
y_test_pneumonia_ = []
y_test_lumpy_virus_ = []
for i in range(len(y_test_anthrax)):
  if(y_test_anthrax[i]):
    # print(y_test_anthrax[i])
    y_test_anthrax_.append([1,0])
  else:
    y_test_anthrax_.append([0,1])

  if(y_test_lumpy_virus[i]):
    # print(y_test_lumpy_virus[i])
    y_test_lumpy_virus_.append([1,0])
  else:
    y_test_lumpy_virus_.append([0,1])

  if(y_test_pneumonia[i]):
    # print(y_test_pneumonia[i])
    y_test_pneumonia_.append([1,0])
  else:
    y_test_pneumonia_.append([0,1])

  if(y_test_blackleg[i]):
    # print(y_test_blackleg[i])
    y_test_blackleg_.append([1,0])
  else:
    y_test_blackleg_.append([0,1])

  if(y_test_food_and_mouth_disease[i]):
    # print(y_test_blackleg[i])
    y_test_food_and_mouth_disease_.append([1,0])
  else:
    y_test_food_and_mouth_disease_.append([0,1])

y_test_anthrax = np.array(y_test_anthrax_)
y_test_blackleg = np.array(y_test_blackleg_)
y_test_pneumonia = np.array(y_test_pneumonia_)
y_test_lumpy_virus = np.array(y_test_lumpy_virus_)
y_test_foot_and_mouth = np.array(y_test_food_and_mouth_disease_)

model_anthrax = Sequential()
model_anthrax.add(Dense(500, activation='relu', input_dim=26))
model_anthrax.add(Dense(100, activation='relu'))
model_anthrax.add(Dense(50, activation='relu'))
model_anthrax.add(Dense(2, activation='softmax'))

# # Compile the model
model_anthrax.compile(optimizer='adam', loss='categorical_crossentropy', metrics=['accuracy'])

model_anthrax.fit(x_test, y_test_anthrax, epochs=5)

# t = np.array(t).reshape(1,26)
t1 = np.array(x_test[8000]).reshape(1,26)
t2 = np.array(x_test[30]).reshape(1,26)
t3 = t1.copy();
for i in range(len(t1[0])):
  if(t1[0][i]):
     t3[0][1] = 1

t3[0][7] =   (1 - t3[0][7])
t3[0][8] =   (1 - t3[0][8])
t3[0][9] =   (1 - t3[0][9])
t3[0][2] =  (1 -  t3[0][10])

print(t3)
pred_test = model_anthrax.predict(t1)
print(pred_test)
# t[0][21] = 1
# t[0][14] = 1
# t[0][8] = 1
# t[0][12] = 1
# t[0][6] = 1
# print(t)
pred_test = model_anthrax.predict(t2)
print(pred_test)
pred_test = model_anthrax.predict(t3)
print(pred_test)

model_blackleg = Sequential()
model_blackleg.add(Dense(500, activation='relu', input_dim=26))
model_blackleg.add(Dense(100, activation='relu'))
model_blackleg.add(Dense(50, activation='relu'))
model_blackleg.add(Dense(2, activation='softmax'))
# Compblackleg model
model_blackleg.compile(optimizer='adam', loss='categorical_crossentropy', metrics=['accuracy'])

model_blackleg.fit(x_test, y_test_blackleg, epochs=5)

model_pneumonia = Sequential()
model_pneumonia.add(Dense(500, activation='relu', input_dim=26))
model_pneumonia.add(Dense(100, activation='relu'))
model_pneumonia.add(Dense(50, activation='relu'))
model_pneumonia.add(Dense(2, activation='softmax'))
# Comppneumonia model
model_pneumonia.compile(optimizer='adam', loss='categorical_crossentropy', metrics=['accuracy'])

model_pneumonia.fit(x_test, y_test_pneumonia, epochs=5)

model_lumpy_virus = Sequential()
model_lumpy_virus.add(Dense(500, activation='relu', input_dim=26))
model_lumpy_virus.add(Dense(100, activation='relu'))
model_lumpy_virus.add(Dense(50, activation='relu'))
model_lumpy_virus.add(Dense(2, activation='softmax'))
# Complumpy_virus model
model_lumpy_virus.compile(optimizer='adam', loss='categorical_crossentropy', metrics=['accuracy'])

model_lumpy_virus.fit(x_test, y_test_lumpy_virus, epochs=5)

model_foot_and_mouth = Sequential()
model_foot_and_mouth.add(Dense(500, activation='relu', input_dim=26))
model_foot_and_mouth.add(Dense(100, activation='relu'))
model_foot_and_mouth.add(Dense(50, activation='relu'))
model_foot_and_mouth.add(Dense(2, activation='softmax'))
# Compfoot_and_mouth model
model_foot_and_mouth.compile(optimizer='adam', loss='categorical_crossentropy', metrics=['accuracy'])

model_foot_and_mouth.fit(x_test, y_test_foot_and_mouth, epochs=5)
df=pd.read_csv("train.csv")
df2=pd.read_csv("test.csv")
print(df.head)

df.head()
train_y=list(df["prognosis"])
test_y=list(df2["prognosis"])
train_x = df.drop("prognosis", axis=1)
test_x = df2.drop("prognosis", axis=1)
train_x.head()

rf_classifier = RandomForestClassifier(n_estimators=100, random_state=93)
rf_classifier.fit(train_x, train_y)

out=rf_classifier.predict(test_x)
data=[]
@app.route('/', methods=['POST'])
def index():
    false=False
    true=True
    listdis = [
    "mastitis",
    "blackleg",
    "bloat",
    "coccidiosis",
    "cryptosporidiosis",
    "displaced_abomasum",
    "gut worms",
    "listeriosis",
    "liver fluke",
    "necrotic_enteritis",
    "peri weaning diarrhoea",
    "rift valley fever",
    "rumen acidosis",
    "traumatic_reticulitis",
    "calf diphtheria",
    "foot rot",
    "foot and mouth",
    "ragwort poisoning",
    "wooden tongue",
    "infectious bovine rhinotracheitis",
    "acetonaemia",
    "fatty liver syndrome",
    "calf pneumonia",
    "schmallen berg virus",
    "trypanosomosis",
    "fog fever"
]
    
    data=eval(request.data)["data"]
    data2=eval(request.data)["data2"]
    input1=[eval(data2[0])]
    other=rf_classifier.predict(input1)
    indexAnimal=listdis.index(other)
    outPneumonia=model_blackleg.predict(data).tolist()
    outLumpy=model_lumpy_virus.predict(data).tolist()
    outFoot=model_foot_and_mouth.predict(data).tolist()
    outBlack=model_blackleg.predict(data).tolist()
    outAnthrax=model_anthrax.predict(data).tolist()
    print(outAnthrax)

    response= jsonify(message=str('{ "Anthhrax" :'+str(outAnthrax)+','+'"Other" :'+str(indexAnimal) +","+'"Lumpy" :'+str(outLumpy)+","+'"BlackLeg":'+str(outBlack)+','+'"Pneunomonia" :'+str(outPneumonia)+","+'"Foot" :'+str(outFoot)+"}"),
                                  other=str(other))
    print(response)
   
    return response.data

if __name__ == '__main__':
    app.run()